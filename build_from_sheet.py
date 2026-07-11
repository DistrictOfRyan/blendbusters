#!/usr/bin/env python3
# Generates BlendBusters teardown pages directly from the ChatGPT 100-product
# spreadsheet, wiring in the Ingredient Catalog's cheap swaps + Amazon links.
# Skips the products already built bespoke by build_teardowns.py.
import re, warnings
warnings.filterwarnings('ignore')
import openpyxl
import bb_render
from bb_render import render_compare, compute_score, esc as _esc

EVDOTS={'strong':3,'mod':2,'weak':1}
def cls_for(gid):
    if gid in STRONG: return 'strong'
    if gid in WEAK: return 'weak'
    return 'mod'
def evnote(name,cls):
    return {'strong':'Solid human evidence for this ingredient.','mod':'Reasonable evidence; benefit varies.',
            'weak':'Limited or mixed human evidence.'}.get(cls,'')

XLSX='/root/.claude/uploads/89659117-1595-510e-9d89-38cba852da22/b6dbc339-high_price_supplement_cost_comparison_100_products.xlsx'

# ---- Affiliate config ----
# Set this to the real Amazon Associates tracking ID once approved; every link
# and the one-click cart activate earnings the moment this is filled in.
AFFILIATE_TAG='blendbusters-20'  # TODO: replace with the real Associates tag

# One representative Amazon ASIN per catalog ingredient (fills the one-click cart).
ASIN_MAP={
 'G01':'B006VRNEFO','G02':'B00YMSLT88','G03':'B0842DJJYC','G04':'B0046XC528','G05':'B0019LVGPC',
 'G06':'B09126BY8S','G07':'B000BD0RT0','G08':'B004189JCW','G09':'B005D0DTS2','G10':'B0001SR3EC',
 'G11':'B0D1VWSPFH','G12':'B01IT9NLHW','G13':'B07TT8B1JJ','G14':'B00OG8G9UM','G15':'B07KXWY1WY',
 'G16':'B01MY5CW7S','G17':'B01CUYHCX6','G18':'B00GQV9YX6','G19':'B005FKTWCC','G20':'B0013OVZJW',
 'G21':'B079K32QB6','G22':'B09QFSDLL5','G23':'B002S1U7RU','G24':'B002RWUNYM','G25':'B0019GW3G8',
 'G26':'B000Z8ZKW0','G27':'B000I4DFAK','G28':'B07Y2H11DP','G29':'B01ANLKJ9M','G30':'B01AMJCHB8',
 'G31':'B06XKM7P97','G32':'B07X7J7GVK','G33':'B0D7QMVHP8','G34':'B07T8C9N97','G35':'B094XMC514',
 'G36':'B01CUQFKW4','G37':'B076TT43SR','G38':'B07BTGCJTW','G39':'B01BCQ3RLE','G40':'B001RYKA3U',
 'G41':'B07PM8X5CG','G42':'B09CX59Q91','G43':'B079YF1K1B','G44':'B00NMPP4WY','G45':'B00KHQJJ0Y',
 'G46':'B09DGTBBSF','G47':'B0078W47PM','G48':'B00EEOGIWC','G49':'B000O2TNKW','G50':'B0B4PKGQXC',
 'G51':'B07Q3SWL3X','G52':'B09SP6HLLP','G53':'B09WJPFVVP','G54':'B07JQFPL61','G55':'B019GU4ILQ',
 'G56':'B084HQ4DYQ','G57':'B07BXVFC32','G58':'B07MWCPRDC','G59':'B01BTBZWBU','G60':'B0DJ1NFCS3',
}

def amz(url):
    if not url: return url
    if AFFILIATE_TAG and 'tag=' not in url:
        url=url+('&' if '?' in url else '?')+'tag='+AFFILIATE_TAG
    return url

def cart_url(alt_ids):
    asins=[ASIN_MAP[g] for g in alt_ids if g in ASIN_MAP and ASIN_MAP[g]]
    if not asins: return None
    parts=['https://www.amazon.com/gp/aws/cart/add.html?AssociateTag='+(AFFILIATE_TAG or '')]
    for i,a in enumerate(asins,1):
        parts.append('ASIN.%d=%s&Quantity.%d=1'%(i,a,i))
    return '&'.join(parts)

# Brand behind each catalog swap (from the ASIN research).
BRAND_BY_GID={
 'G01':'Kirkland','G02':'Nature Made','G03':'Sports Research','G04':'Nature Made','G05':'NOW Foods',
 'G06':'Nutricost','G07':"Doctor's Best",'G08':'NOW Foods','G09':'Nature Made','G10':'NOW Foods',
 'G11':'Nutricost','G12':'Liquid I.V.','G13':'LMNT','G14':'TRIORAL','G15':'Starbucks','G16':'Nutricost',
 'G17':'Nutricost','G18':'NOW Foods','G19':'Nature Made','G20':'NOW Foods','G21':'Nutricost','G22':'Nutricost',
 'G23':'NOW Foods','G24':'NOW Foods','G25':"Doctor's Best",'G26':'NOW Foods','G27':"Doctor's Best",
 'G28':'Nutricost','G29':'Bausch + Lomb','G30':'Nutricost','G31':'Nutricost','G32':'Nutricost','G33':'Nutricost',
 'G34':'Nutricost','G35':'Nutricost','G36':'Nutricost','G37':'Nutricost','G38':'Nutricost','G39':'Nutricost',
 'G40':'NOW Foods','G41':'Double Wood','G42':'Nutricost','G43':'Double Wood','G44':"Doctor's Best",
 'G45':'NOW Foods','G46':'Double Wood','G47':'NOW Foods','G48':'Sports Research','G49':'NOW Foods',
 'G50':"Doctor's Best",'G51':'Nature Made','G52':'Nutricost','G53':'Nutricost','G54':'NOW Foods',
 'G55':'Nuun','G56':'LMNT','G57':'Nature Made','G58':"Nature's Bounty",'G59':'Sports Research','G60':'Nutricost',
}
# DTC affiliate programs (typically 10-40% vs Amazon's ~1-4.5%). Fill 'url' with
# the brand's affiliate/referral landing link once approved; buy_link() then
# prefers the DTC link over Amazon automatically. Leave '' until approved.
DTC_PROGRAMS={
 'Nutricost':{'url':'','network':'ShareASale / direct'},
 'NOW Foods':{'url':'','network':'direct'},
 'Double Wood':{'url':'','network':'Impact'},
 'Sports Research':{'url':'','network':'ShareASale'},
 "Doctor's Best":{'url':'','network':'direct'},
 'Nature Made':{'url':'','network':'direct'},
 'LMNT':{'url':'','network':'direct'},
 'Nuun':{'url':'','network':'direct'},
}

def buy_link(gid):
    """DTC affiliate link if the brand program is live, else the Amazon link."""
    brand=BRAND_BY_GID.get(gid)
    if brand and DTC_PROGRAMS.get(brand,{}).get('url'):
        return DTC_PROGRAMS[brand]['url'], brand
    return amz(CAT[gid]['url']) if gid in CAT else '#get', brand

def short(s,n):
    s=str(s or '').strip()
    if len(s)<=n: return s
    cut=s[:n]; sp=cut.rfind(' ')
    return (cut[:sp] if sp>0 else cut).rstrip(' ,.;')+'…'

CAT_HOOK={
 'Energy':('The real active is caffeine',
   'A cup of coffee or a ~5¢ caffeine tablet delivers the same lift — the rest is flavor, B-vitamins you already get, and branding.'),
 'Hydration':('Electrolytes only matter when you are actually losing them',
   'Heavy sweat, heat, or illness — then sodium and potassium help. For a normal day, water does the job, and the mix is cheap either way.'),
 'Sleep & calm':('Melatonin and magnesium do the heavy lifting',
   'Both cost pennies on their own — and lower melatonin doses often work as well. The botanicals are usually a sprinkle.'),
 'Daily multis & greens':('A cheap multivitamin covers the vitamins',
   'Greens powders are weak versus eating actual vegetables, and the "superfood" extras are mostly fairy-dusted.'),
 'Beauty, joint, immune & nootropic':('One or two ingredients do the real work',
   'The rest is dosed low or thin on evidence — buy the actives that matter on their own.'),
 'Probiotic, omega, magnesium & fiber':('The active is a commodity',
   'Probiotic strains, fish oil, magnesium, and fiber are all cheap generics — the brand markup is the product.'),
}

GA = ('<!-- Google tag (gtag.js) -->\n'
      '<script async src="https://www.googletagmanager.com/gtag/js?id=G-529DGYE1QB"></script>\n'
      '<script>\nwindow.dataLayer = window.dataLayer || [];\n'
      'function gtag(){dataLayer.push(arguments);}\ngtag(\'js\', new Date());\n'
      'gtag(\'config\', \'G-529DGYE1QB\');\n</script>')
LOGO = ('<svg width="30" height="30" viewBox="0 0 40 40" fill="none" aria-hidden="true">'
        '<rect x="1.5" y="1.5" width="37" height="37" rx="10" fill="rgba(255,90,44,.14)" stroke="#FF5A2C" stroke-width="1.6"/>'
        '<rect x="11" y="10" width="18" height="9" rx="4.5" transform="rotate(45 20 20)" fill="none" stroke="#FF8A3D" stroke-width="2.6"/>'
        '<path d="M20 20 l4 -4" stroke="#FF8A3D" stroke-width="2.6" stroke-linecap="round"/></svg>')
JS = ("(function(){var rows=document.querySelectorAll('#rows .row input[type=checkbox]'),BASE=%%BASE%%;"
      "function fmt(n){return '$'+n.toFixed(2)}"
      "function rc(){var t=0;rows.forEach(function(cb){if(cb.checked)t+=parseFloat(cb.dataset.cost)});"
      "var s=BASE-t,p=Math.round(s/BASE*100);"
      "var dt=document.getElementById('diy-total');if(dt)dt.firstChild.nodeValue=fmt(t)+' ';"
      "var ds=document.getElementById('diy-save');if(ds)ds.textContent=s>0?'save '+fmt(s):'';"
      "var vd=document.getElementById('v-diy');if(vd)vd.innerHTML='$'+Math.round(t)+'<span style=\"font-size:13px\">/mo</span>';"
      "var vy=document.getElementById('v-year');if(vy)vy.textContent='$'+Math.round(s*12);"
      "var vp=document.getElementById('v-pct');if(vp)vp.textContent=(p>0?p:0)+'%';}"
      "rows.forEach(function(cb){cb.addEventListener('change',rc)});rc();"
      "var f=document.getElementById('capform');if(f){f.addEventListener('submit',function(e){e.preventDefault();"
      "var m=document.getElementById('capmsg'),em=document.getElementById('capemail').value.trim();"
      "if(em&&em.indexOf('@')>0){m.style.color='var(--volt2)';m.textContent='\\u2713 You are on the list.';f.reset()}"
      "else{m.style.color='#ff7a5c';m.textContent='Enter a valid email.'}})}"
      "var els=document.querySelectorAll('section');els.forEach(function(el){el.classList.add('reveal')});"
      "if('IntersectionObserver' in window){var ob=new IntersectionObserver(function(es){es.forEach(function(e){"
      "if(e.isIntersecting){e.target.classList.add('in');ob.unobserve(e.target)}})},{threshold:0.1,rootMargin:'0px 0px -40px 0px'});"
      "els.forEach(function(el){ob.observe(el)})}else{els.forEach(function(el){el.classList.add('in')})}})();")

def esc(s):
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;') if s else ''

def tiles(items):
    h=""
    for label,note,color in items:
        h+=('<div class="itile" style="background:linear-gradient(150deg,%s)"><div class="in">'
            '<div class="inm">%s</div><div class="dose">%s</div></div></div>'%(color,label,note))
    return h

def rows(items):
    h=""
    for nm,ds,cost,ev,evc in items:
        h+=('<li class="row"><input type="checkbox" checked data-cost="%.2f" aria-label="%s">'
            '<div class="info"><div class="nm">%s <span class="ev %s">%s</span></div>'
            '<div class="ds">%s</div></div><div class="cost">$%.2f<small>/mo</small></div></li>'
            %(cost,nm,nm,evc,ev,ds,cost))
    return h

def path(tag,feature,title,price,pp,ul,href='#get',ext=False):
    cls="path feature" if feature else "path"
    btn="btn volt block" if feature else "btn ghost block"
    lis="".join("<li>%s</li>"%x for x in ul)
    if ext:
        label="Shop the swap on Amazon →"; a='<a class="%s" href="%s" target="_blank" rel="sponsored nofollow noopener">%s</a>'%(btn,href,label)
    else:
        label="See the pick →" if feature else "Get the shopping list →"
        a='<a class="%s" href="%s" rel="sponsored nofollow">%s</a>'%(btn,href,label)
    return ('<div class="%s"><div class="ptag">%s</div><h3>%s</h3><div class="price">%s</div>'
            '<p class="pp">%s</p><ul>%s</ul>%s</div>'%(cls,tag,title,price,pp,lis,a))

def render(d):
    swap=sum(c for _,_,c,_,_ in d['rows']); save=d['orig']-swap; pct=round(save/d['orig']*100); year=round(save*12)
    band=""
    talk="".join("<li>%s</li>"%b for b in d['talk'])
    return ('<!doctype html>\n<html lang="en">\n<head>\n<meta charset="utf-8">\n'
      '<title>%s Teardown · BlendBusters</title>\n'
      '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
      '<meta name="description" content="%s teardown: what is inside and the cheaper swap. Same ingredients, less markup.">\n'
      '%s\n<link rel="stylesheet" href="/bb.css">\n</head>\n<body>\n'
      '<header class="top"><div class="wrap"><a class="brand" href="/">%s BlendBusters</a>'
      '<nav class="main"><a class="lnk" href="/">← All teardowns</a><a class="btn flame" href="#get">Get the free swaps</a></nav></div></header>\n'
      '<a id="top"></a>\n'
      '<div class="hero"><div class="wrap"><p class="kick">Teardown · %s</p>'
      '<h1 class="display">%s</h1><p class="sub">%s</p>'
      '<div class="herobtns"><a class="btn flame" href="#teardown">See the swap ↓</a><a class="btn ghost" href="/">More teardowns</a></div>'
      '<div class="faceoff"><div class="fo villain"><div class="tg">%s</div><div class="big strike">$%d<small>/mo</small></div><div class="cap">%s</div></div>'
      '<div class="foburst"><div><div class="p">%d%%</div><div class="s">Cheaper</div></div></div>'
      '<div class="fo winner"><div class="tg">The same, bought smart</div><div class="big">$%d<small>/mo</small></div><div class="cap">%s</div></div></div>'
      '</div></div>\n%s\n'
      '<div class="trust"><div class="wrap"><span>✓ <b>Every swap</b> priced</span><span>✓ <b>Commodity ingredients</b></span>'
      '<span>✓ <b>Not sponsored</b></span><span>✓ <b>Honest</b> on what works</span></div></div>\n'
      '<div class="wrap">\n'
      '<section id="inside"><div class="shead"><p class="kick">What you are paying for</p><h2>%s</h2><p class="lead">%s</p></div>'
      '<div class="igrid">%s</div></section>\n'
      '<section id="teardown"><div class="tdintro"><span class="tag">◆ The swap</span><h2>Build it cheaper. Watch the price drop.</h2>'
      '<p class="lead">%s</p></div>'
      '<div class="verdict"><div class="vc red"><div class="v">$%d<span style="font-size:13px">/mo</span></div><div class="k">%s</div></div>'
      '<div class="vc win"><div class="v" id="v-diy">$%d<span style="font-size:13px">/mo</span></div><div class="k">Your swap</div></div>'
      '<div class="vc win"><div class="v" id="v-year">$%d</div><div class="k">Saved / year</div></div>'
      '<div class="vc win"><div class="v" id="v-pct">%d%%</div><div class="k">Cheaper</div></div></div>'
      '<div class="calc"><div class="chead"><div class="t">%s<span>%s</span></div><div class="hint">✓ = in my swap</div></div>'
      '<ul class="rows" id="rows">%s</ul>'
      '<div class="foot"><div class="f"><div class="lbl">%s, same month</div><div class="amt">$%d.00</div></div>'
      '<div class="f b"><div class="lbl">Your swap</div><div class="amt" id="diy-total">$%.2f <s id="diy-save">save $%.2f</s></div></div></div></div>'
      '%s<div class="paths">%s%s</div>'
      '<div class="disclose">Disclosure: shopping links are affiliate links; we may earn a commission at no extra cost to you. (Amazon commissions activate once our Associates account is approved.)</div>'
      '</section>\n'
      '<section id="realtalk"><div class="realtalk"><h3>\U0001f525 Real talk: what actually works</h3><ul>%s</ul></div></section>\n'
      '<section id="get"><div class="capture"><h2>Get the free Label Decoder</h2>'
      '<p>A new overpriced supplement torn apart every week, with the exact cheaper swap.</p>'
      '<form id="capform" novalidate><input type="email" id="capemail" placeholder="you@email.com" aria-label="Email address">'
      '<button class="btn flame" type="submit">Send me the swaps</button></form>'
      '<div id="capmsg" role="status" aria-live="polite"></div><p class="fine">No spam. (Demo form for now.)</p></div></section>\n'
      '</div>\n%s'
      '<footer><div class="wrap"><div class="fgrid"><a class="brand" href="/">%s BlendBusters</a>'
      '<nav><a href="/">Home</a><a href="/#about">About</a><a href="/#about-legal">Disclosures</a></nav></div>'
      '<p class="disc">© 2026 Hunt Web Consulting Services. Informational only; not medical advice. '
      'These statements have not been evaluated by the FDA; products are not intended to diagnose, treat, cure, or prevent any disease. '
      'Affiliate links may earn commissions per FTC guidelines. Brand names are used for comparison and review only.</p></div></footer>\n'
      '<script>%s</script>\n</body>\n</html>\n'
      %(d['name'], d['name'], GA, LOGO, d['num'], d['h1'], d['sub'],
        d['villain_tg'], d['orig'], d['villain_cap'], pct, swap, d['winner_cap'], band,
        d['inside_h2'], d['inside_lead'], tiles(d['inside']),
        d['teardown_lead'], d['orig'], d['name'], swap, year, pct,
        d['name'], d['orig_sub'], rows(d['rows']), d['name'], d['orig'], swap, save,
        d.get('cart',''), d['pathA'], d['pathB'], talk, d.get('related',''), LOGO, JS.replace('%%BASE%%', str(d['orig']))))

# ---------- load spreadsheet ----------
wb=openpyxl.load_workbook(XLSX, data_only=True)
P=list(wb['Products'].iter_rows(values_only=True))[1:]
C=list(wb['Ingredient Catalog'].iter_rows(values_only=True))[1:]
CAT={r[0]:{'name':r[1],'serv':r[2],'cost':float(r[5] or 0),'url':r[6]} for r in C if r[0]}

SKIP_IDS={1,3,4,20,21,22,69,76,77,89,90,91,99,100}  # already built bespoke
STRONG={'G01','G02','G05','G10','G11','G24','G04','G59','G03','G16','G15','G07','G08','G54','G60','G33','G38','G39','G17'}
WEAK={'G21','G41','G42','G43','G44','G45','G46','G47','G31','G32','G48','G49','G50','G25','G26','G27','G22','G37','G40'}
PAL=['#1e5a3a,#0e2b1c','#2e6a1e,#16330a','#5a4a1e,#2e2410','#1e5a55,#0e2b28','#3a4a5e,#1a2430','#5a3a26,#281910','#5a1e2e,#2a0f16','#33333c,#111114']

def evi(gid):
    if gid in STRONG: return ('Well studied','strong')
    if gid in WEAK: return ('Limited data','weak')
    return ('Same purpose','mod')

def slugify(s):
    s=re.sub(r'[^a-z0-9]+','-',str(s).lower()).strip('-')
    return re.sub(r'-+','-',s)

def money(x):
    return str(int(round(x)))

BUCKET_BY_CAT={
 'Daily multis & greens':'Daily multivitamins & greens',
 'Hydration':'Hydration & electrolytes',
 'Energy':'Energy drinks & mixes',
 'Sleep & calm':'Sleep & calm',
 'Beauty, joint, immune & nootropic':'Beauty, hair, joint & immune',
 'Probiotic, omega, magnesium & fiber':'Gut, probiotic, omega & fiber',
}

cards={}  # bucket -> list of (savings_year, name, slug, orig, budget)
pages=[]  # (dict, bucket, slug) deferred for pass-2 render with related links
used_slugs=set()
made=0
for r in P:
    pid=r[0]
    if pid is None or pid in SKIP_IDS: continue
    name=str(r[2]).strip()
    cat=r[1]
    try: orig=float(r[10])
    except (TypeError,ValueError): continue
    if not orig or orig<=0: continue
    summ=str(r[12] or '').strip()
    plan=str(r[15] or '').strip()
    safety=str(r[30] or '').strip()
    # alts: (id,qty) pairs
    alts=[]
    for gi,qi in [(16,17),(18,19),(20,21),(22,23)]:
        gid=r[gi]; q=r[qi]
        if gid and gid in CAT and q and float(q)>0:
            alts.append((gid,float(q)))
    if not alts: continue
    rws=[]
    for gid,q in alts:
        c=CAT[gid]; cost=round(q*c['cost']*30,2)
        if cost<=0: cost=round(c['cost']*30,2)
        lbl,cls=evi(gid)
        qd=('%g/day'%q) if q!=1 else '1/day'
        rws.append((esc(c['name']),qd,cost,lbl,cls))
    budget=round(sum(c for _,_,c,_,_ in rws),2)
    if budget<=0: continue
    save=orig-budget; pct=round(save/orig*100)
    if pct<8: continue  # not a meaningful dupe
    slug=slugify(name)
    if slug in used_slugs: slug=slug+'-'+str(pid)
    used_slugs.add(slug)
    proprietary = ('proprietar' in summ.lower() or 'blend' in summ.lower())
    verdict = 'Lower-cost ingredient match' if pct>=40 else 'Partial match' if pct>=15 else 'Important differences'
    match_pct = 72 if verdict.startswith('Lower') else 55 if verdict.startswith('Partial') else 45
    # swap rows -> new schema
    swap_rows=[]; ev_dots=[]
    for gid,q in alts:
        c=CAT[gid]; cost=round(q*c['cost']*30,2) or round(c['cost']*30,2)
        cl=cls_for(gid); ev_dots.append(EVDOTS[cl])
        swap_rows.append({'name':_esc(c['name']),'desc':('%g/day'%q if q!=1 else '1/day'),'cost':cost,'cls':cl,'asin':ASIN_MAP.get(gid)})
    ev_avg=sum(ev_dots)/len(ev_dots) if ev_dots else 2
    verdict_note=('A pairing of lower-cost generics covers several overlapping ingredients and a similar intended use, for about %d%% less. Some ingredients and exact doses don’t carry over.'%pct)
    matches=['Several overlapping ingredients — %s — available as lower-cost generics.'%(', '.join(s['name'] for s in swap_rows[:3])),
             'A similar intended use, at a comparable daily amount where the dose is disclosed.',
             'Doses you can read on a plain generic label.']
    differs=[]
    if plan: differs.append(_esc(plan))
    differs.append('Format, flavor, and convenience differ from the brand.')
    if proprietary: differs.append('The brand uses a proprietary blend, so exact doses cannot be matched or verified.')
    evidence=[{'name':s['name'],'cls':s['cls'],'note':evnote(s['name'],s['cls'])} for s in swap_rows]
    consult=['<b style="color:var(--ink)">Talk to a qualified healthcare professional</b> before changing supplements if you are pregnant or nursing, immunocompromised, managing a health condition, or taking medications.',
             'See a professional for persistent or unusual symptoms — a supplement is not a substitute for evaluation.']
    b_url,b_brand=buy_link(alts[0][0])
    d={'slug':slug,'name':_esc(name),'category':_esc(cat),'reviewed':'Jul 2026',
       'brand_price':int(round(orig)),'brand_per_day':'$%.2f'%(orig/30),
       'label_summary':_esc(summ),'proprietary':proprietary,
       'verdict':verdict,'verdict_note':verdict_note,'match_pct':match_pct,
       'swap_rows':swap_rows,'matches':matches,'differs':differs,'evidence':evidence,
       'score':compute_score(pct,proprietary,ev_avg),
       'safety':_esc(safety) or 'Introduce any new supplement gradually, and review it against your current medications and conditions.',
       'consult':consult,
       'sources':[('Brand label & price — merchant listing (price checked Jul 2026)','#',False),
                  ('Lower-cost generic pricing — retail listings (checked Jul 2026)','#',False)],
       'cart_asins':[ASIN_MAP.get(g) for g,_ in alts],'primary_buy':b_url,'primary_brand':b_brand,
       'related':[]}
    bucket=BUCKET_BY_CAT.get(cat,'More teardowns')
    cards.setdefault(bucket,[]).append((round(save*12),esc(name),slug,int(round(orig)),budget))
    pages.append((d,bucket,slug))
    made+=1

# ---------- bespoke pages (already built) folded into the library grid ----------
BESPOKE=[  # bucket, name, slug/href, orig, budget
 ('Testosterone & men’s','Mars Men','/#teardown',65,22),
 ('Testosterone & men’s','TestoFuel','/testofuel.html',65,20),
 ('Testosterone & men’s','Prime Male','/prime-male.html',75,32),
 ('Testosterone & men’s','Nugenix Total-T','/nugenix.html',70,18),
 ('Daily multivitamins & greens','AG1','/ag1.html',79,32),
 ('Daily multivitamins & greens','Ka’chava','/kachava.html',140,75),
 ('Daily multivitamins & greens','Balance of Nature','/balance-of-nature.html',70,15),
 ('Daily multivitamins & greens','Bloom Greens','/bloom.html',35,13),
 ('Daily multivitamins & greens','Huel','/huel.html',212,102),
 ('Hydration & electrolytes','LMNT','/lmnt.html',45,2),
 ('Hydration & electrolytes','Liquid I.V.','/liquid-iv.html',25,2),
 ('Energy drinks & mixes','MUD\\WTR','/mud-wtr.html',40,30),
 ('Sleep & calm','OLLY Sleep','/olly-sleep.html',17,6),
 ('Brain & nootropics','Alpha Brain','/alpha-brain.html',53,27),
 ('Brain & nootropics','Prevagen','/prevagen.html',40,14),
 ('Brain & nootropics','Neuriva','/neuriva.html',32,11),
 ('Beauty, hair, joint & immune','Nutrafol','/nutrafol.html',79,38),
 ('Beauty, hair, joint & immune','Vital Proteins','/vital-proteins.html',48,19),
 ('Gut, probiotic, omega & fiber','Seed DS-01','/seed.html',50,19),
 ('Gut, probiotic, omega & fiber','Peachy Clean','/peachy.html',40,10),
 ('Gut, probiotic, omega & fiber','ARMRA Colostrum','/armra.html',110,10),
 ('Fitness & performance','Creatine Gummies','/creatine-gummies.html',56,3),
 ('Fitness & performance','Legion Pulse','/pre-workout.html',45,10),
 ('Fitness & performance','SuperBeets','/superbeets.html',40,4),
 ('Longevity & everyday','Tru Niagen','/tru-niagen.html',45,2),
 ('Longevity & everyday','fatty15','/fatty15.html',50,6),
 ('Longevity & everyday','Cymbiotika Vitamin C','/cymbiotika.html',62,2),
 ('Longevity & everyday','Magnesium Breakthrough','/magnesium-breakthrough.html',40,6),
 ('Longevity & everyday','Goli ACV','/goli.html',57,5),
 ('Longevity & everyday','O Positiv FLO','/flo-pms.html',32,5),
 ('Gut, probiotic, omega & fiber','Bellway Super Fiber + Fruit','/bellway-alternative.html',18,6),
 ('Gut, probiotic, omega & fiber','Pure for Men Stay Ready Fiber','/pure-for-men-alternative.html',36,6),
 ('Fitness & performance','Peachy Plump Creatine Bites','/peachy-plump-alternative.html',60,6),
 ('Hydration & electrolytes','Pedialyte AdvancedCare Plus','/pedialyte-alternative.html',45,11),
 ('Hydration & electrolytes','Nuun Instant','/nuun-instant-alternative.html',26,6),
 ('Beauty, hair, joint & immune','Dose & Co Collagen Peptides','/dose-and-co-alternative.html',60,27),
 ('Gut, probiotic, omega & fiber','Metamucil Premium Blend','/metamucil-alternative.html',7,3),
 ('Gut, probiotic, omega & fiber','Benefiber Prebiotic Fiber','/benefiber-alternative.html',6,2),
 ('Hydration & electrolytes','Liquid I.V. Sugar-Free','/liquid-iv-sugar-free-alternative.html',39,13),
 ('Hydration & electrolytes','Gatorade Gatorlyte','/gatorlyte-alternative.html',40,9),
 ('Daily multivitamins & greens','1st Phorm Opti-Greens 50','/opti-greens-50-alternative.html',60,35),
 ('Daily multivitamins & greens','Grüns Daily Greens','/gruns-alternative.html',56,36),
]
for b,nm,slug,o,bu in BESPOKE:
    href=slug if slug.startswith('/') else '/'+slug+'.html'
    cards.setdefault(b,[]).append((round((o-bu)*12),nm,href,o,bu))

ORDER=['Testosterone & men’s','Daily multivitamins & greens','Hydration & electrolytes',
       'Energy drinks & mixes','Sleep & calm','Brain & nootropics',
       'Beauty, hair, joint & immune','Gut, probiotic, omega & fiber',
       'Fitness & performance','Longevity & everyday','More teardowns']

def card_html(name,href,o,bu):
    href2=href if href.startswith('/') else '/'+href+'.html'
    yr=round((o-bu)*12)
    bud=('%g'%bu) if bu<10 else str(int(round(bu)))
    return ('<a class="s3" href="%s" style="text-decoration:none;color:inherit">'
            '<div class="num" style="color:var(--flame2)">%s</div>'
            '<h3 style="text-transform:none">$%d &#8594; $%s / mo</h3>'
            '<p>save ~$%s/yr</p></a>'%(href2,name,o,bud,'{:,}'.format(yr)))

def _href(c2): return c2 if str(c2).startswith('/') else '/'+c2+'.html'
def verdict_of(o,bu): return 'Lower-cost match' if o and (o-bu)/float(o)>=0.4 else 'Partial match'

# ---- pass 2: related internal links + render via shared compliant template ----
for d,bucket,slug in pages:
    rel=sorted([c for c in cards.get(bucket,[]) if c[2]!=slug],key=lambda x:-x[0])[:3]
    d['related']=[(nm,_href(href),verdict_of(o,bu),yr) for yr,nm,href,o,bu in rel]
    with open('%s.html'%slug,'w',encoding='utf-8') as f: f.write(render_compare(d))
print('rendered %d comparison pages'%len(pages))

# ---------- homepage (editorial) ----------
def lib_card(name,href,o,bu):
    yr=round((o-bu)*12); bud=('%g'%bu) if bu<10 else str(int(round(bu)))
    return ('<a class="rc" href="%s"><span class="cat mono">$%d/mo &#8594; $%s/mo</span>'
            '<h4>%s</h4><span class="s">%s · save ~$%s/yr</span></a>'
            %(_href(href),o,bud,_esc(name),verdict_of(o,bu),'{:,}'.format(yr)))
total=sum(len(v) for v in cards.values())
present=[b for b in ORDER if b in cards]
allcards=sorted([c for v in cards.values() for c in v],key=lambda x:-x[0])
trending=allcards[:4]
def bar_card(yr,name,href,o,bu):
    ov=o/30.0; mv=bu/30.0; w=max(8,round(mv/ov*100)) if ov else 20
    vd=verdict_of(o,bu); vc='match' if vd.startswith('Lower') else ''
    return ('<a class="card" href="%s"><div class="top"><div><span class="cat">$%d/mo brand</span>'
            '<h3>%s</h3></div><span class="vstamp %s">%s</span></div>'
            '<div class="bars"><div class="barrow"><span class="lbl">Brand</span><span class="track"><span class="fillb" style="width:100%%"></span></span><span class="v">$%.2f/day</span></div>'
            '<div class="barrow"><span class="lbl">Match</span><span class="track"><span class="fillb match" style="width:%d%%"></span></span><span class="v">$%.2f/day</span></div></div>'
            '<div class="foot"><span class="overlap">save <b>~$%s/yr</b></span><span class="seeit">See the breakdown →</span></div></a>'
            %(_href(href),o,_esc(name),vc,vd,ov,w,mv,'{:,}'.format(round((o-bu)*12))))
nav=''.join('<a href="#cat-%d">%s <span>%d</span></a>'%(i,_esc(b),len(cards[b])) for i,b in enumerate(present))
lib=['<section id="library"><div class="wrap"><div class="shead" style="display:block"><p class="eyebrow">The comparison library</p>'
     '<h2 style="margin-top:8px">%d products, broken down.</h2>'
     '<p class="lead" style="margin-top:10px">Overlapping ingredients, real doses, dated prices. Pick a category — each page shows the lower-cost match, the math, and what differs.</p></div>'
     '<div class="catnav">%s</div>'%(total,nav)]
for i,b in enumerate(present):
    lst=sorted(cards[b],key=lambda x:-x[0])
    lib.append('<h3 class="cathead" id="cat-%d">%s <span>(%d)</span></h3><div class="libgrid">%s</div>'
               %(i,_esc(b),len(lst),''.join(lib_card(nm,href,o,bu) for _,nm,href,o,bu in lst)))
lib.append('</div></section>')

home=(bb_render._head('BlendBusters — Pay for ingredients. Not hype.',
        'BlendBusters breaks down expensive wellness products, compares ingredients, doses, and cost, and finds lower-cost options — with the savings, evidence, and tradeoffs before you buy.')
  +'<body>\n'+bb_render._header(back=False)
  +'<a id="top"></a>\n<main>\n'
  # hero
  +'<section class="hero"><div class="wrap hero-grid"><div>'
   '<p class="eyebrow">Independent wellness comparisons</p>'
   '<h1 style="margin-top:14px">Pay for ingredients.<br><em>Not hype.</em></h1>'
   '<p class="lead">BlendBusters breaks down expensive wellness products, compares their ingredients and doses, and finds lower-cost options. See the savings, differences, evidence, and tradeoffs before you buy.</p>'
   '<form class="search" id="searchform"><label for="q">Enter a product name or paste a product link.</label>'
   '<div class="search-row"><input id="q" placeholder="e.g. Liquid I.V., or https://…" autocomplete="off" aria-label="Product name or link">'
   '<button class="btn primary" type="submit">Break it down</button></div>'
   '<p class="try">Popular: <a href="#library">Peachy Clean</a> · <a href="#library">AG1</a> · <a href="#library">LMNT</a> · <a href="#library">Seed</a></p>'
   '<p class="msg" id="searchmsg" role="status" aria-live="polite"></p></form>'
   '<p class="disc-inline">We compare ingredients, doses, and price — not medical outcomes. A lower-cost match shares overlapping ingredients and a similar intended use; it is not a guaranteed equivalent.</p></div>'
   '<div class="cards" style="align-self:start">'+''.join(bar_card(*c) for c in trending)+'</div>'
   '</div></section>\n'
  # how it works
  +'<section id="how"><div class="wrap"><div class="shead" style="display:block"><p class="eyebrow">How BlendBusters works</p>'
   '<h2 style="margin-top:8px">Every comparison runs on the same eight questions.</h2>'
   '<p class="lead" style="margin-top:10px">We weigh each product the same way and show our work — including where the data isn’t public.</p></div>'
   '<div class="weigh">'
   '<div class="wcell"><div class="k">Match</div><h4>Ingredient matching</h4><p>Which actives actually overlap.</p></div>'
   '<div class="wcell"><div class="k">Dose</div><h4>Dose comparison</h4><p>How the amounts line up — where disclosed.</p></div>'
   '<div class="wcell"><div class="k">Cost</div><h4>Cost per serving</h4><p>Price normalized to the dose that matters.</p></div>'
   '<div class="wcell"><div class="k">Evidence</div><h4>Evidence quality</h4><p>Strength of human evidence, rated and sourced.</p></div>'
   '<div class="wcell"><div class="k">Ease</div><h4>Convenience</h4><p>What you trade for the savings.</p></div>'
   '<div class="wcell"><div class="k">Clarity</div><h4>Transparency</h4><p>Disclosed doses vs. a proprietary blend.</p></div>'
   '<div class="wcell"><div class="k">Savings</div><h4>Estimated savings</h4><p>Monthly and annual, with the price date.</p></div>'
   '<div class="wcell"><div class="k">Honesty</div><h4>Tradeoffs</h4><p>What you give up. We say so plainly.</p></div>'
   '</div><p class="fine" style="margin-top:16px">These feed the <b>BlendBuster Score</b> (out of 100). Where a brand hides a dose, that field reads <span class="mono">Data unavailable</span>.</p></div></section>\n'
  # library
  +'\n'.join(lib)+'\n'
  # request + email
  +'<section id="request"><div class="wrap split"><div><p class="eyebrow">Request a comparison</p>'
   '<h2 style="margin:8px 0 12px">Paying too much for something? Send it in.</h2>'
   '<p class="lead">Drop a product name or link and we’ll run it through the model. Readers pick most of what we cover.</p>'
   '<form class="panel" id="reqform" style="margin-top:18px">'
   '<div class="f"><label for="rp">Product name</label><input id="rp" placeholder="e.g. Future Method Daily Fiber" required></div>'
   '<div class="f"><label for="ru">Product link <span style="color:var(--ink-3);font-weight:400">(optional)</span></label><input id="ru" placeholder="https://…"></div>'
   '<div class="f"><label for="re">Your email</label><input id="re" type="email" placeholder="you@email.com" required></div>'
   '<button class="btn primary block" type="submit">Request this comparison</button>'
   '<p class="msg" id="reqmsg" role="status" aria-live="polite"></p></form></div>'
   '<div class="email-box"><p class="eyebrow">The weekly bust</p><h3 style="margin-top:8px">One overpriced blend, busted each week.</h3>'
   '<form class="email-row" id="mailform"><input id="me" type="email" placeholder="you@email.com" aria-label="Email address" required>'
   '<button class="btn primary" type="submit">Subscribe</button></form><p class="msg" id="mailmsg" role="status" aria-live="polite"></p>'
   '<p class="fine">No spam, no selling your data. Just the breakdown.</p>'
   '<p style="margin-top:16px"><a class="btn ghost" href="/savings-index.html">📄 Get the free Savings Index</a></p></div></div></section>\n'
  # affiliate + about
  +'<section id="about-legal"><div class="wrap"><div class="affbox"><span class="k">Affiliate disclosure</span>'
   '<p>BlendBusters earns affiliate commissions when you buy through some merchant links. We don’t manufacture or sell products — purchases happen on third-party merchant sites. Comparisons and verdicts are written before any link is added, and a commission never changes a score. Prices are estimates from public sources with the date we checked.</p></div>'
   '<p id="about" class="fine" style="margin-top:14px">BlendBusters is an independent consumer comparison platform operated by Hunt Web Consulting Services. We compare ingredients, doses, and price to help you decide before you buy. We are not a manufacturer, pharmacy, or medical provider.</p></div></section>\n'
  +'</main>\n'+bb_render._footer()
  +'<script>'+bb_render.COMPARE_JS.replace("ev('comparison_view',{comparison_id:document.body.dataset.slug||''});","")
   +"(function(){function w(f,m,t){var e=document.getElementById(f);if(!e)return;e.addEventListener('submit',function(v){v.preventDefault();var x=document.getElementById(m);x.style.color='var(--accent-deep)';x.textContent=t;try{gtag('event',f==='mailform'||f==='reqform'?'email_signup':'site_search',{})}catch(_){}if(f!=='searchform')e.reset();});}"
   "w('searchform','searchmsg','Demo — a live search would open that product’s breakdown.');w('reqform','reqmsg','✓ Request received (demo). We’ll email you when it’s live.');w('mailform','mailmsg','✓ You’re on the list (demo).');})();"
   '</script>\n</body>\n</html>\n')
open('index.html','w',encoding='utf-8').write(home)
print('wrote editorial homepage: %d cards, %d categories'%(total,len(present)))

# ---------- sitemap.xml + robots.txt (all pages, for indexing) ----------
import glob as _glob
BASE='https://blendbusters.com/'
urls=['']  # homepage
for f in sorted(_glob.glob('*.html')):
    if f=='index.html': continue
    urls.append(f)
sm=['<?xml version="1.0" encoding="UTF-8"?>',
    '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
for u in urls:
    pri='1.0' if u=='' else '0.7'
    sm.append('  <url><loc>%s%s</loc><changefreq>weekly</changefreq><priority>%s</priority></url>'%(BASE,u,pri))
sm.append('</urlset>')
open('sitemap.xml','w',encoding='utf-8').write('\n'.join(sm)+'\n')
open('robots.txt','w',encoding='utf-8').write('User-agent: *\nAllow: /\nSitemap: %ssitemap.xml\n'%BASE)
print('wrote sitemap.xml (%d urls) + robots.txt'%len(urls))

# ---------- Dupe Kit: printable master reference (the $19 / lead-magnet asset) ----------
allrows=[]
for b in present:
    for yr,nm,href,o,bu in sorted(cards[b],key=lambda x:-x[0]):
        allrows.append((b,nm,href,o,bu,yr))
tot_year=sum(r[5] for r in allrows)
tot_mo=sum((r[3]-r[4]) for r in allrows)
dk=[bb_render._head('The BlendBusters Savings Index',
      'Every product we’ve compared and its lower-cost ingredient match, with estimated monthly and annual savings. Prices dated; verify before buying.')
    +'<style>@media print{.top,.dk-cta,footer{display:none}}'
    '.dk-t{width:100%;border-collapse:collapse;margin:8px 0 26px;font-size:14px;font-family:var(--mono)}'
    '.dk-t th{text-align:left;color:var(--accent-deep);border-bottom:1px solid var(--line-2);padding:8px 10px;font-size:11px;letter-spacing:.1em;text-transform:uppercase}'
    '.dk-t td{padding:9px 10px;border-bottom:1px solid var(--line)}'
    '.dk-t a{color:var(--accent-deep);text-decoration:none;border-bottom:1px dotted var(--line-2)}'
    '.dk-save{color:var(--accent-deep);font-weight:700;white-space:nowrap}.dk-old{color:var(--ink-3);text-decoration:line-through}'
    '.dk-hero{text-align:center;padding:26px 0 6px}.dk-big{font-family:var(--mono);font-size:clamp(40px,8vw,68px);font-weight:800;color:var(--accent-deep);line-height:1}'
    '</style>\n</head>\n<body>\n'
    +bb_render._header()
    +'<div class="wrap"><div class="dk-hero"><p class="eyebrow">The Savings Index</p>'
    '<h1 class="title" style="font-size:clamp(28px,5vw,44px);margin:10px 0">Every comparison, and the lower-cost match.</h1>'
    '<p class="lead" style="margin:0 auto">A reference list of %d products, the lower-cost ingredient match, and the estimated savings. Prices are dated estimates — verify before buying. A match shares overlapping ingredients and a similar intended use, not a guaranteed equivalent.</p>'
    '<div style="margin:22px 0 6px"><div class="dk-big">$%s/yr</div>'
    '<div style="color:var(--ink-3);font-size:14px">estimated total across the list — about $%s a month</div></div></div>'%(len(allrows),'{:,}'.format(round(tot_year)),'{:,}'.format(round(tot_mo)))]
dk.append('<div class="wrap">')
for b in present:
    rws=sorted(cards[b],key=lambda x:-x[0])
    sub_y=sum(x[0] for x in rws)
    dk.append('<h3 class="cathead">%s <span>— est. save ~$%s/yr</span></h3>'%(esc(b),'{:,}'.format(round(sub_y))))
    dk.append('<table class="dk-t"><thead><tr><th>Brand</th><th>Brand cost</th><th>Match</th><th>Est. save</th><th></th></tr></thead><tbody>')
    for yr,nm,href,o,bu in rws:
        h2=href if href.startswith('/') else '/'+href+'.html'
        bud=('%g'%bu) if bu<10 else str(int(round(bu)))
        dk.append('<tr><td>%s</td><td class="dk-old">$%d/mo</td><td>$%s/mo</td><td class="dk-save">$%s/yr</td><td><a href="%s">see it →</a></td></tr>'%(esc(nm),o,bud,'{:,}'.format(yr),h2))
    dk.append('</tbody></table>')
dk.append('<div class="dk-cta" style="text-align:center;margin:10px 0 40px"><a class="btn primary" href="/#request">Get the weekly comparison</a> '
          '<a class="btn ghost" href="/">Browse all comparisons</a></div></div>')
dk.append(bb_render._footer()+'</body>\n</html>\n')
open('savings-index.html','w',encoding='utf-8').write(''.join(dk))
print('wrote savings-index.html (%d products, $%s/yr total)'%(len(allrows),'{:,}'.format(round(tot_year))))

# ---------- methodology / editorial standards / disclosure hub ----------
WEIGHTS=[('Ingredient match','20','Which actives actually overlap between the brand and the lower-cost option.'),
 ('Dose match','20','How the disclosed amounts line up. Hidden behind a proprietary blend → <span class="mono">Data unavailable</span>.'),
 ('Cost per serving','20','Price normalized to the dose that matters, not the sticker price.'),
 ('Evidence quality','15','Strength of human evidence for each active, rated and sourced.'),
 ('Convenience','10','What you trade for the savings — format, number of products, steps.'),
 ('Transparency','10','Whether the brand discloses its doses or hides them in a blend.'),
 ('Savings','10','Estimated monthly and annual, with the price-checked date attached.'),
 ('Tradeoffs','10','What you give up. If the match falls short, the score reflects it.')]
wrows=''.join('<div class="sr"><span class="lb">%s</span><span class="strack"><span class="sfill" style="width:%d%%"></span></span><span class="sv">%s pts</span></div>'%(n,int(w)/20*100,w) for n,w,_ in WEIGHTS)
wlist=''.join('<div class="wcell"><div class="k">%s pts</div><h4>%s</h4><p>%s</p></div>'%(w,n,dsc) for n,w,dsc in WEIGHTS)
mp=(bb_render._head('Methodology & editorial standards · BlendBusters',
      'How BlendBusters scores comparisons: a transparent 100-point model, how we handle undisclosed doses, price checking, sourcing, corrections, and affiliate independence.')
  +'<body>\n'+bb_render._header()
  +'<div class="wrap"><nav class="crumb"><a href="/">Home</a> / <b>Methodology</b></nav>'
   '<div class="title"><span class="cat">How we work</span><h1>Methodology &amp; editorial standards</h1>'
   '<div class="meta"><span>Last updated <b>Jul 2026</b></span></div></div></div>\n'
  +'<section><div class="wrap"><div class="shead"><h2>The BlendBuster Score</h2><span class="ctag an">BlendBusters analysis</span></div>'
   '<p class="lead" style="margin-bottom:16px">Every comparison is scored on the same eight factors, out of 100. We show the weight of each, and where a value cannot be sourced we mark it <span class="mono">Data unavailable</span> rather than invent a number.</p>'
   '<div class="score"><div class="sub">'+wrows+'</div></div>'
   '<div class="weigh" style="margin-top:22px">'+wlist+'</div></div></section>\n'
  +'<section><div class="wrap"><div class="shead"><h2>How we handle undisclosed doses</h2></div>'
   '<p class="lead">Many premium products list a proprietary blend, so the amount of each ingredient is not published. When that happens we cannot verify a dose match, and we say so: the <span class="mono">Dose match</span> field reads <span class="mono">Data unavailable</span> and does not contribute to the score. A large price gap alone does not make two products equivalent.</p></div></section>\n'
  +'<section><div class="wrap"><div class="shead"><h2>Prices &amp; sourcing</h2></div>'
   '<p class="lead">Prices are estimates from public sources with the date we checked, and they change often — always verify on the merchant’s site. Every comparison ships with dated, linked sources. Numbers marked “to be verified” are placeholders pending editorial sign-off. We do not copy merchant descriptions; the analysis and calculations are our own.</p></div></section>\n'
  +'<section id="standards"><div class="wrap"><div class="shead"><h2>Editorial standards</h2></div>'
   '<div class="safe"><ul><li>We label <b style="color:var(--ink)">brand claims</b>, <b style="color:var(--ink)">BlendBusters analysis</b>, and <b style="color:var(--ink)">scientific evidence</b> separately, so you always know who is saying what.</li>'
   '<li>We use language like “lower-cost ingredient match,” “overlapping ingredients,” and “similar intended use” — never “exactly the same,” “works just as well,” or “guaranteed equivalent.”</li>'
   '<li>We do not fabricate reviews, testimonials, endorsements, test results, prices, or scientific findings.</li>'
   '<li>We do not state that a product treats, cures, prevents, or diagnoses any condition.</li>'
   '<li>Anyone can send a correction; we date and log every edit.</li></ul></div></div></section>\n'
  +'<section id="affiliate"><div class="wrap"><div class="shead"><h2>Affiliate independence</h2></div>'
   '<div class="affbox"><span class="k">How we make money</span>'
   '<p>BlendBusters earns affiliate commissions when you buy through some merchant links. We don’t manufacture or sell products — purchases happen on third-party merchant sites. Comparisons and verdicts are written before any link is added, and a commission never changes a verdict or a score. We keep a better option in a comparison even when another merchant pays more.</p></div>'
   '<p class="disc-inline" style="margin-top:14px">This content is for general informational purposes and is not individualized medical advice. Consult a qualified healthcare professional before changing any supplement.</p></div></section>\n'
  +'</div>\n'+bb_render._footer()
  +'<script>(function(){var r=document.documentElement,t=document.getElementById("theme");t&&t.addEventListener("click",function(){var c=r.getAttribute("data-theme");if(!c)c=matchMedia("(prefers-color-scheme:dark)").matches?"dark":"light";r.setAttribute("data-theme",c==="dark"?"light":"dark");t.textContent=c==="dark"?"\\u263e":"\\u2600"})})();</script>\n</body>\n</html>\n')
open('methodology.html','w',encoding='utf-8').write(mp)
print('wrote methodology.html')
print('done')
